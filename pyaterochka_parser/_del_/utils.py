import os
import json
import pandas as pd
from datetime import datetime
import requests
from config import TelegramBot
import openpyxl
from config import db_retail_map   # ← путь к Excel-файлу из config.py



# === Функции записи Excel и JSON ===

def write_excel(data, filename, stats=None):
    """
    Записывает список словарей (data) в Excel.
    """
    if not data:
        print("⚠️ Нет данных для записи в Excel.")
        return
    try:
        df = pd.DataFrame(data)
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        df.to_excel(filename, index=False)

        print(f"✅ Файл Excel записан успешно: {filename}")
        if stats:
            print(f"📊 Статистика: {stats}")
    except Exception as e:
        print(f"❌ Ошибка при записи Excel: {e}")


def write_json(data, filename):
    """
    Записывает данные в JSON-файл.
    """
    try:
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"💾 JSON сохранён: {filename}")
    except Exception as e:
        print(f"❌ Ошибка записи JSON: {e}")


# === Функция логирования ===

def write_log(text, filename="parser_log.txt"):
    """
    Пишет текст в лог-файл с меткой времени.
    """
    try:
        os.makedirs("logs", exist_ok=True)
        filepath = os.path.join("logs", filename)
        with open(filepath, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {text}\n")
        print(f"📝 Лог записан: {text}")
    except Exception as e:
        print(f"⚠️ Ошибка записи в лог: {e}")

# === Унифицированная функция логирования по ТЗ ===

#def write_log_uni(text, filename=None):
    """
    Универсальная запись лога.
    Формат строки:
    YYYY-MM-DD \t HH:MM \t <text>
    """
"""
    # 1) Проверка: имя файла передано?
    if not filename or str(filename).strip() == "":
        print("⚠️ Параметр 'filename' пуст. Лог не записан.")
        return False

    try:
        # 2) Создаём каталог log при необходимости
        log_dir = "log"
        os.makedirs(log_dir, exist_ok=True)

        # 3) Формируем корректный путь (Только logs + filename)
        filepath = os.path.join(log_dir, filename)

        # 4) Запись строки
        date_str = datetime.now().strftime("%Y-%m-%d")
        time_str = datetime.now().strftime("%H:%M")

        with open(filepath, "a", encoding="utf-8") as f:
            f.write(f"{date_str}\t{time_str}\t🚀{my_id}\t{text}\n")

        print(f"📝 Записано в лог: {text}")
        return True

    except Exception as e:
        print(f"⚠️ Ошибка записи в лог: {e}")
    
    return False
"""
# === Основная функция отправки уведомлений в Telegram ===
def SendTelegram(message: str):
    """
    Отправляет сообщение в Telegram.
    Использует TOKEN, CHAT_ID и send из config.TelegramBot.
    Поддерживает HTML и эмодзи.
    """
    send_flag = TelegramBot.get("send", False)
    if not send_flag:
        print("📴 Отправка в Telegram отключена (send=False).")
        return False

    token = TelegramBot.get("TOKEN")
    chat_id = TelegramBot.get("CHAT_ID")
    my_id   = TelegramBot.get("MY_ID")	

    if not token or not chat_id:
        print("⚠️ Ошибка: TOKEN или CHAT_ID не указаны в config.TelegramBot")
        return False

    # ✅ Формируем сообщение с нормальными переводами строк
    message = f"🚀 {my_id} 🚀\n{message}"

    url = f"https://api.telegram.org/bot{token}/sendMessage"
    data = {
        "chat_id": chat_id,
        "text": message,
        "parse_mode": "HTML"
    }

    try:
        response = requests.post(url, data=data, timeout=10)
        if response.status_code == 200:
            print("💬 Сообщение успешно отправлено в Telegram.")
            return True
        else:
            print(f"⚠️ Ошибка Telegram API: {response.status_code} {response.text}")
            return False
    except Exception as e:
        print(f"❌ Ошибка отправки Telegram: {e}")
        return False


# === Дополнительная функция: отправка файла в Telegram ===

def SendTelegramFile(filepath: str, caption: str = ""):
    """
    Отправляет файл (Excel, лог, JSON и т.д.) в Telegram.
    Использует параметры из config.TelegramBot.
    """
    send_flag = TelegramBot.get("send", False)
    if not send_flag:
        print("📴 Отправка файлов в Telegram отключена (send=False).")
        return False

    token = TelegramBot.get("TOKEN")
    chat_id = TelegramBot.get("CHAT_ID")
    my_id   = TelegramBot.get("MY_ID")	


    if not token or not chat_id:
        print("⚠️ Ошибка: TOKEN или CHAT_ID не указаны в config.TelegramBot")
        return False

    if not os.path.exists(filepath):
        print(f"⚠️ Файл не найден: {filepath}")
        return False

    caption = f"🚀 {my_id} 🚀\n{caption}"

    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with open(filepath, "rb") as file_data:
        files = {"document": file_data}
        data = {"chat_id": chat_id, "caption": caption}

        try:
            response = requests.post(url, data=data, files=files, timeout=30)
            if response.status_code == 200:
                print(f"📤 Файл {os.path.basename(filepath)} отправлен в Telegram.")
                return True
            else:
                print(f"⚠️ Ошибка Telegram API: {response.status_code} {response.text}")
                return False
        except Exception as e:
            print(f"❌ Ошибка при отправке файла в Telegram: {e}")
            return False

def update_retail_points(parser_name: str, city: str, count: int) -> bool:
    """
    Обновляет количество торговых точек и фиксирует событие в Excel-файле.

    СТРУКТУРА ВЫХОДНОГО ФАЙЛА (строго):
        A: Дата
        B: Время
        C: Имя парсера
        D: Город
        E: Количество торговых точек

    ЛОГИКА РАБОТЫ:
        1. Путь к файлу берём ТОЛЬКО из config.py → db_retail_map.
        2. Если файл указан без .xlsx → добавляем расширение.
        3. Если файл отсутствует → создаём новый с правильной шапкой.
        4. Ищем строку с parser_name + city.
        5. Если строка найдена → обновляем только если новое count > старого.
        6. Если строка не найдена → создаём новую.
        7. Фиксируем дату и время обновления.
        8. Возвращаем True, если данные обновлены; False — если нет изменений.
    """

    # --- 1. Корректируем путь: добавляем .xlsx, если забыли ---
    filename = db_retail_map
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    # --- 2. Если файла нет — создаём новый с шапкой ---
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Дата", "Время", "Имя парсера", "Город", "Количество ТТ"])
        wb.save(filename)
    else:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

    updated = False
    today = datetime.now().strftime("%Y-%m-%d")
    now_time = datetime.now().strftime("%H:%M:%S")

    found_row = None

    # --- 3. Ищем строку, где совпадают парсер и город ---
    for row in ws.iter_rows(min_row=2):
        p = str(row[2].value).strip().lower() if row[2].value else ""
        c = str(row[3].value).strip().lower() if row[3].value else ""

        if p == parser_name.lower() and c == city.lower():
            found_row = row
            break

    # --- 4. Если строка найдена → обновляем при условии роста count ---
    if found_row:
        old_count = found_row[4].value

        # обновляем только если новое значение больше старого
        if old_count is None or count > int(old_count):
            found_row[0].value = today
            found_row[1].value = now_time
            found_row[2].value = parser_name
            found_row[3].value = city
            found_row[4].value = count
            updated = True

    # --- 5. Если строки нет — создаём новую ---
    else:
        ws.append([today, now_time, parser_name, city, count])
        updated = True

    # --- 6. Сохраняем только если были изменения ---
    if updated:
        wb.save(filename)
        print(f"➡️ Обновлено: {parser_name} — {city} количество ТТ → {count} ")				#		ТТ  ({today} {now_time})
    else:
        print(f"ℹ️ Без изменений: {parser_name} — {city}  количество ТТ <= {count} ")			#		(старое значение больше или равно)")

    return updated
